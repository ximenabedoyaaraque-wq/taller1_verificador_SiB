"""
Taller 1 · Verificador de Localidades SiB Colombia
Protocolo SiB Colombia / Instituto Humboldt — Método radio-punto
Colecciones Biológicas Universidad CES
"""

import re
import unicodedata
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# ── Configuración de la app ────────────────────────────────────────
st.set_page_config(
    page_title="Taller 1 · Verificador de Localidades SiB Colombia",
    page_icon="🗺️",
    layout="centered"
)

st.markdown("""
<style>
    .titulo { font-size: 1.6rem; font-weight: 700; color: #1F6B40; }
    .subtitulo { font-size: 1rem; color: #555; margin-bottom: 1.5rem; }
    .instruccion { background: #E8F5EE; padding: 1rem; border-radius: 8px;
                   border-left: 4px solid #1F6B40; margin-bottom: 1rem; }
    .advertencia { background: #FFF9E6; padding: 1rem; border-radius: 8px;
                   border-left: 4px solid #E8A020; margin-bottom: 1rem; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="titulo">🗺️ Taller 1 · Verificador de Localidades SiB Colombia</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitulo">Protocolo de georreferenciación SiB Colombia · Colecciones Biológicas Universidad CES</div>', unsafe_allow_html=True)

st.markdown("""
<div class="instruccion">
<b>¿Cómo usar esta herramienta?</b><br>
1. Sube tu archivo Excel o CSV con los registros biológicos.<br>
2. Haz clic en <b>Analizar</b>.<br>
3. Descarga el reporte con los errores marcados.<br>
4. Usa el reporte para corregir manualmente en OpenRefine.
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="advertencia">
⚠️ <b>Esta herramienta detecta errores de formato y sintaxis — no corrige automáticamente.</b><br>
Los nombres de localidades requieren ojo crítico: verifica el nombre oficial de cada lugar
antes de hacer cualquier cambio en OpenRefine.
</div>
""", unsafe_allow_html=True)


# ── Columnas esperadas (nombres Darwin Core) ───────────────────────
COLUMNAS_LOCALIDAD = [
    "verbatimlocality", "locality", "verbatimLocality",
    "localidad", "localidad original", "localidad estandarizada"
]
COLUMNAS_PAIS = ["country", "país", "pais"]
COLUMNAS_DEPTO = ["stateprovince", "departamento", "stateProvince"]
COLUMNAS_MUNICIPIO = ["county", "municipio"]

# ── Palabras/siglas que NO deben marcarse como siglas ─────────────
# Palabras en mayúsculas que son válidas o nombres geográficos conocidos
PALABRAS_MAYUSCULAS_VALIDAS = {
    # Palabras técnicas y estados válidos
    "SIN", "DATOS", "NaN", "NA", "N/A",
    # País
    "COLOMBIA",
    # Instituciones colombianas cuyo nombre oficial es una sigla
    # (agregar aquí las siglas válidas de tu colección)
    "CES",       # Universidad CES — válida en bases de datos CBUCES
    "IAVH",      # Instituto Humboldt
    "IDEAM",     # Instituto de Hidrología
    "IGAC",      # Instituto Geográfico Agustín Codazzi
    "SINCHI",    # Instituto Amazónico
    "INDERENA",  # Instituto histórico de recursos naturales
    "UNAL",      # Universidad Nacional de Colombia
    "UDEA",      # Universidad de Antioquia
    "UIS",       # Universidad Industrial de Santander
    "UPTC",      # Universidad Pedagógica y Tecnológica de Colombia
}

# NOTA PARA OTRAS COLECCIONES:
# Si usas este código en una colección diferente a CBUCES,
# revisa esta lista y agrega o quita las siglas institucionales
# que sean válidas en tu contexto. Por ejemplo, si tu colección
# pertenece a la Universidad del Valle, agrega "UNIVALLE".
# Las siglas de ÁREAS PROTEGIDAS nunca son válidas (PNN, SFF, etc.)
# independientemente de la colección.

# ── Siglas de áreas protegidas (para mensaje específico) ──────────
SIGLAS_AREAS_PROTEGIDAS = {
    "PNN", "SFF", "RNA", "DMI", "DRMI", "VPP", "ANU", "RNSC", "ZRC",
}

# ── Palabras con tilde frecuentemente omitidas ───────────────────
PALABRAS_CON_TILDE = {
    "paramo": "Páramo", "paramos": "Páramos",
    "area": "Área", "areas": "Áreas",
    "nucleo": "Núcleo",
    "arbol": "Árbol", "arboles": "Árboles",
    "bano": "Baño",
    "narino": "Nariño",
    "bogota": "Bogotá",
    "medellin": "Medellín",
    "popayan": "Popayán",
    "cucuta": "Cúcuta",
    "monteria": "Montería",
    "quibdo": "Quibdó",
    "mitu": "Mitú",
    "san jose": "San José",
}

# ── Separadores incorrectos ───────────────────────────────────────
SEPARADORES_INCORRECTOS = [';', ':', '...', '…']


def normalizar(texto):
    """Quita tildes y pasa a minúsculas para comparar."""
    if not isinstance(texto, str):
        return ""
    nfkd = unicodedata.normalize('NFKD', texto.lower())
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def detectar_siglas_generales(loc):
    """
    Detecta cualquier palabra escrita completamente en mayúsculas
    de 2 o más letras que no sea una excepción conocida.
    Cubre siglas como: CVZ, RR, CBZS, URR, PNN, SFF, etc.
    """
    errores = []
    # Buscar palabras de 2+ letras todas mayúsculas (incluyendo tildes en mayúsculas)
    patron = r'\b[A-ZÁÉÍÓÚÑ]{2,}\b'
    matches = re.findall(patron, loc)
    for m in matches:
        if m in PALABRAS_MAYUSCULAS_VALIDAS:
            continue
        if m in SIGLAS_AREAS_PROTEGIDAS:
            errores.append(
                f"Sigla de área protegida '{m}' — usar nombre oficial completo "
                f"(ej: Parque Nacional Natural, Santuario de Flora y Fauna...)"
            )
        else:
            errores.append(
                f"Posible sigla o abreviatura en mayúsculas: '{m}' — "
                f"verificar si debe escribirse con el nombre completo"
            )
    return errores


def detectar_abreviaturas_generales(loc):
    """
    Detecta abreviaturas: palabras cortas (2-5 letras) seguidas de punto
    que no sean el final de la cadena ni números.
    Ej: Vda. Hda. Mpio. Dpto. Carr. Cra. etc.
    """
    errores = []
    # Patrón: 2-5 letras seguidas de punto, que no sea fin de texto
    patron = r'\b([A-Za-záéíóúñÁÉÍÓÚÑ]{2,5})\.'
    matches = re.finditer(patron, loc)
    for m in matches:
        palabra = m.group(1)
        # Ignorar si es número romano o si la palabra completa está en el texto sin punto
        errores.append(
            f"Posible abreviatura: '{m.group()}' — verificar si debe escribirse completo"
        )
    return errores


def detectar_errores(localidad):
    """
    Recibe una cadena de localidad y retorna lista de errores encontrados.
    """
    if not isinstance(localidad, str) or localidad.strip() == "":
        return []

    errores = []
    loc = localidad.strip()

    # 1. Minúscula al inicio
    if loc[0].islower():
        errores.append("La descripción debe iniciar con mayúscula")

    # 2. Punto final
    if loc.endswith('.'):
        # Evitar falso positivo si termina en sigla con punto (ya cubierta arriba)
        errores.append("Tiene punto final (no permitido al terminar la localidad)")

    # 3. Separadores incorrectos
    for sep in SEPARADORES_INCORRECTOS:
        if sep in loc:
            errores.append(f"Separador incorrecto '{sep}' — usar solo comas (,)")
            break

    # 4. Siglas en mayúsculas (detección general — incluye áreas protegidas)
    errores.extend(detectar_siglas_generales(loc))

    # 5. Abreviaturas con punto (detección general)
    # Solo si no termina en punto (ya detectado) para no duplicar
    loc_sin_fin = loc.rstrip('.')
    abrev = detectar_abreviaturas_generales(loc_sin_fin)
    errores.extend(abrev)

    # 6. Tildes faltantes
    loc_norm = normalizar(loc)
    for sin_tilde, con_tilde in PALABRAS_CON_TILDE.items():
        if re.search(r'\b' + sin_tilde + r'\b', loc_norm):
            if con_tilde.lower() not in loc.lower():
                errores.append(
                    f"Posible tilde faltante: '{sin_tilde.capitalize()}' "
                    f"→ debería ser '{con_tilde}'"
                )

    # 7. Comillas innecesarias
    if '"' in loc or '\u201c' in loc or '\u201d' in loc:
        errores.append(
            "Uso de comillas — solo se usan para citas textuales o descripciones ambiguas"
        )

    # 8. Conector 'y' entre descripciones (debe ser coma)
    if re.search(r',\s+y\s+', loc):
        errores.append(
            "Posible conector 'y' — verificar si debe reemplazarse por coma"
        )

    # 9. Información de hábitat dentro de localidad
    palabras_habitat = ['bosque', 'pastizal', 'rastrojo', 'potrero', 'cultivo',
                        'páramo', 'subpáramo', 'humedal', 'manglar']
    loc_norm2 = normalizar(loc)
    for hab in palabras_habitat:
        if re.search(r'\b' + normalizar(hab) + r'\b', loc_norm2):
            errores.append(
                f"Posible información de hábitat ('{hab}') dentro del campo localidad — "
                f"no hace parte del campo localidad según el protocolo SiB"
            )
            break

    return errores


def limpiar_nombre_col(nombre):
    """
    Normaliza un nombre de columna para comparación:
    quita asteriscos (*), espacios extra y pasa a minúsculas.
    Necesario porque el estándar Darwin Core marca columnas obligatorias con *
    (ej: '*verbatimLocality' o 'verbatimLocality *').
    """
    return nombre.replace('*', '').strip().lower()


def encontrar_columna(df, opciones):
    """
    Busca una columna en el DataFrame por varias variantes de nombre.
    Ignora asteriscos (*) y espacios extra en los nombres de columnas
    (formato común en plantillas Darwin Core del SiB Colombia).
    """
    cols_limpias = {limpiar_nombre_col(c): c for c in df.columns}
    for opcion in opciones:
        if opcion.lower() in cols_limpias:
            return cols_limpias[opcion.lower()]
    return None


def generar_reporte(df, col_localidad, col_pais=None, col_depto=None, col_municipio=None):
    """Analiza el DataFrame y retorna BytesIO con el Excel de reporte."""
    df_reporte = df.copy()
    errores_por_fila = []

    for _, fila in df.iterrows():
        localidad = fila.get(col_localidad, "") if col_localidad else ""
        errores = detectar_errores(str(localidad) if pd.notna(localidad) else "")

        if col_pais and pd.isna(fila.get(col_pais)):
            if isinstance(localidad, str) and localidad.strip():
                errores.append("Campo País vacío con localidad diligenciada")
        if col_depto and pd.isna(fila.get(col_depto)):
            if isinstance(localidad, str) and localidad.strip():
                errores.append("Campo Departamento vacío con localidad diligenciada")
        if col_municipio and pd.isna(fila.get(col_municipio)):
            if isinstance(localidad, str) and localidad.strip():
                errores.append("Campo Municipio vacío con localidad diligenciada")

        errores_por_fila.append(errores)

    df_reporte["Errores detectados"] = [
        " | ".join(e) if e else "" for e in errores_por_fila
    ]
    df_reporte["Estado"] = [
        "⚠ Con errores" if e else "✓ Sin errores detectados"
        for e in errores_por_fila
    ]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_reporte.to_excel(writer, sheet_name='Registros con errores', index=False)

        conteo = {}
        for lista in errores_por_fila:
            for e in lista:
                tipo = e.split("—")[0].strip()[:70]
                conteo[tipo] = conteo.get(tipo, 0) + 1

        df_resumen = pd.DataFrame({
            "Tipo de error": list(conteo.keys()),
            "Registros afectados": list(conteo.values())
        }).sort_values("Registros afectados", ascending=False)
        df_resumen.to_excel(writer, sheet_name='Resumen de errores', index=False)

        # Formato hoja 1
        ws = writer.sheets['Registros con errores']
        amarillo = PatternFill("solid", fgColor="FFF3CD")
        verde_claro = PatternFill("solid", fgColor="D4EDDA")
        header_fill = PatternFill("solid", fgColor="1F6B40")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        borde = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = borde
        for row in ws.iter_rows(min_row=2):
            tiene_error = "⚠" in str(row[-1].value)
            fill = amarillo if tiene_error else verde_claro
            for cell in row:
                cell.fill = fill
                cell.border = borde
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # Formato hoja resumen
        ws2 = writer.sheets['Resumen de errores']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
        for col_idx, col in enumerate(ws2.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 70)

    output.seek(0)
    return output, df_reporte, conteo


# ── Interfaz principal ─────────────────────────────────────────────
archivo = st.file_uploader(
    "Sube tu archivo Excel (.xlsx) o CSV",
    type=["xlsx", "csv"],
    help="Debe contener al menos una columna de localidad (verbatimLocality, locality o localidad)"
)

if archivo is not None:
    try:
        if archivo.name.endswith('.csv'):
            df = pd.read_csv(archivo, encoding='utf-8-sig', low_memory=False)
        else:
            df = pd.read_excel(archivo)

        st.success(f"✓ Archivo cargado: **{archivo.name}** — {len(df)} registros, {len(df.columns)} columnas")

        col_localidad = encontrar_columna(df, COLUMNAS_LOCALIDAD)
        col_pais = encontrar_columna(df, COLUMNAS_PAIS)
        col_depto = encontrar_columna(df, COLUMNAS_DEPTO)
        col_municipio = encontrar_columna(df, COLUMNAS_MUNICIPIO)

        if col_localidad is None:
            st.error(
                "⛔ No encontré columna de localidad. "
                "Asegúrate de que tu archivo tenga una columna llamada "
                "'verbatimLocality', 'locality', 'localidad' o 'localidad original'."
            )
        else:
            st.info(f"📋 Columna de localidad detectada: **{col_localidad}**")
            if col_pais:
                st.caption(
                    f"País: {col_pais} | "
                    f"Departamento: {col_depto or 'no encontrado'} | "
                    f"Municipio: {col_municipio or 'no encontrado'}"
                )

            if st.button("🔍 Analizar errores", type="primary", use_container_width=True):
                with st.spinner("Analizando localidades según el protocolo SiB..."):
                    output, df_reporte, conteo = generar_reporte(
                        df, col_localidad, col_pais, col_depto, col_municipio
                    )

                total = len(df_reporte)
                con_errores = df_reporte["Estado"].str.contains("⚠").sum()
                sin_errores = total - con_errores

                st.markdown("---")
                col1, col2, col3 = st.columns(3)
                col1.metric("Total registros", total)
                col2.metric("Con errores", con_errores,
                            delta=f"{con_errores/total*100:.0f}%", delta_color="inverse")
                col3.metric("Sin errores detectados", sin_errores)

                if conteo:
                    st.markdown("**Errores más frecuentes:**")
                    for tipo, cantidad in sorted(conteo.items(), key=lambda x: -x[1])[:8]:
                        st.markdown(f"- {tipo}: **{cantidad}** registro(s)")

                nombre_reporte = (
                    archivo.name.replace('.xlsx', '').replace('.csv', '')
                    + "_reporte_errores.xlsx"
                )
                st.download_button(
                    label="📥 Descargar reporte Excel",
                    data=output,
                    file_name=nombre_reporte,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )

                st.markdown("""
<div class="instruccion">
<b>¿Qué hacer con el reporte?</b><br>
1. Abre el reporte — las filas <b>amarillas</b> tienen errores detectados.<br>
2. Lee la columna <b>"Errores detectados"</b> para cada registro.<br>
3. Verifica el nombre oficial del lugar (Google Maps, Divipola, RUNAP).<br>
4. Corrige en <b>OpenRefine</b>: doble clic → Editar → nombre correcto.<br>
5. Exporta el CSV corregido.<br><br>
⚠️ <b>Recuerda:</b> esta herramienta detecta errores de formato y sintaxis.
Los nombres geográficos mal escritos pero sin errores de formato
requieren revisión manual y criterio curatorial.
</div>
""", unsafe_allow_html=True)

    except Exception as e:
        st.error(f"⛔ Error al leer el archivo: {e}")
        st.caption("Verifica que el archivo no esté protegido con contraseña y que sea Excel o CSV válido.")

st.markdown("---")
st.caption("Colecciones Biológicas · Universidad CES · Protocolo SiB Colombia / Instituto Humboldt")
